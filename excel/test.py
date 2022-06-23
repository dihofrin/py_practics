import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_excel('supermarket_sales.xlsx')  # reading excel file
excel_file[['Gender', 'Product line', 'Total']]  # choosing specific columns

report_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(
    0)  # creating table

report_table.to_excel('report_2021.xlsx', sheet_name='Report', startrow=4)  # exporting table to excel file

# creating link to row and column
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# cell references(original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# adding diagrams to excel (using module Barchart from openpyxl)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']

# barchart
barchart = BarChart()

# locate data and categories
data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row,
                 max_row=max_row)  # including headers

categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1,
                       max_row=max_row)  # not including headers

# adding data and categories

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# location chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5 # choose the chart style

sheet['B7'] = '=SUM(B5:B6)'
sheet['B7'].style = 'Currency'
wb.save('report_2021.xlsx')

